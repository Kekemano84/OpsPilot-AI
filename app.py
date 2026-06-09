import os
import secrets
import json
import sqlite3
import uuid
import smtplib
import ssl
from datetime import datetime, timedelta
from functools import wraps
from email.message import EmailMessage
from io import BytesIO

from flask import (
    Flask, render_template, request, redirect, url_for, send_file,
    flash, session
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

try:
    import stripe
except Exception:
    stripe = None

try:
    import requests
except Exception:
    requests = None


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "opspilot-ai-dev-secret")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
DB_PATH = os.path.join(DATA_DIR, "opspilot_ai.db")

HMRC_MILE_RATE = 0.45
PERSONAL_ALLOWANCE = 12570
BASIC_RATE_LIMIT = 50270
INCOME_TAX_BASIC = 0.20
INCOME_TAX_HIGHER = 0.40
CLASS2_NI_WEEKLY = 3.45
CLASS4_NI_LOWER = 12570
CLASS4_NI_UPPER = 50270
CLASS4_NI_BASIC = 0.06
CLASS4_NI_HIGHER = 0.02

PLAN_ORDER = {"free": 0, "pro": 1, "business": 2}
PLAN_NAMES = {"free": "Free", "pro": "Pro", "business": "Pro"}
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
STRIPE_SECRET_KEY = os.environ.get("STRIPE_SECRET_KEY")
STRIPE_PRO_PRICE_ID = os.environ.get("STRIPE_PRO_PRICE_ID")
STRIPE_BUSINESS_PRICE_ID = os.environ.get("STRIPE_BUSINESS_PRICE_ID")
APP_BASE_URL = os.environ.get("APP_BASE_URL", "http://127.0.0.1:5000")
SMTP_HOST = os.environ.get("SMTP_HOST")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD")
SMTP_FROM = os.environ.get("SMTP_FROM", SMTP_USER or "")
GOOGLE_MAPS_API_KEY = os.environ.get("GOOGLE_MAPS_API_KEY")


def get_db():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS system_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            plan TEXT NOT NULL DEFAULT 'free',
            business_name TEXT,
            phone TEXT,
            address TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS mileage (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            start_location TEXT NOT NULL,
            end_location TEXT NOT NULL,
            miles REAL NOT NULL,
            purpose TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            category TEXT NOT NULL,
            description TEXT NOT NULL,
            amount REAL NOT NULL,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            invoice_number TEXT NOT NULL,
            date TEXT NOT NULL,
            customer_name TEXT NOT NULL,
            customer_email TEXT,
            description TEXT NOT NULL,
            amount REAL NOT NULL,
            status TEXT NOT NULL DEFAULT 'Unpaid',
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS yard_checks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            trailer_id TEXT NOT NULL,
            location_type TEXT NOT NULL,
            location_detail TEXT,
            status TEXT NOT NULL DEFAULT 'Recorded',
            notes TEXT,
            source TEXT NOT NULL DEFAULT 'Manual',
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS kpi_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            shift TEXT NOT NULL,
            volume INTEGER DEFAULT 0,
            planned_hc INTEGER DEFAULT 0,
            actual_hc INTEGER DEFAULT 0,
            target_rate REAL DEFAULT 0,
            actual_rate REAL DEFAULT 0,
            late_trailers INTEGER DEFAULT 0,
            errors INTEGER DEFAULT 0,
            notes TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS handovers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            shift TEXT NOT NULL,
            manager TEXT,
            volume INTEGER DEFAULT 0,
            planned_hc INTEGER DEFAULT 0,
            actual_hc INTEGER DEFAULT 0,
            late_trailers INTEGER DEFAULT 0,
            issues TEXT,
            actions TEXT,
            generated_report TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS team_members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            role TEXT NOT NULL,
            email TEXT,
            status TEXT NOT NULL DEFAULT 'Active',
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS shift_plans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            shift TEXT NOT NULL,
            volume INTEGER DEFAULT 0,
            available_hc INTEGER DEFAULT 0,
            target_rate REAL DEFAULT 0,
            planned_hours REAL DEFAULT 0,
            ai_plan TEXT,
            created_at TEXT NOT NULL
        )
    """)


    cur.execute("""
        CREATE TABLE IF NOT EXISTS shift_calendar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'Off',
            shift_name TEXT,
            start_time TEXT,
            end_time TEXT,
            notes TEXT,
            source TEXT DEFAULT 'Generated',
            created_at TEXT NOT NULL,
            UNIQUE(user_id, date)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS morning_briefs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            shift TEXT NOT NULL,
            manager TEXT,
            role TEXT,
            volume INTEGER DEFAULT 0,
            available_hc INTEGER DEFAULT 0,
            late_trailers INTEGER DEFAULT 0,
            safety_message TEXT,
            priorities TEXT,
            team_messages TEXT,
            break_reminder TEXT,
            equipment_reminder TEXT,
            generated_brief TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS photo_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            image_filename TEXT,
            trailer_id TEXT,
            location_detail TEXT,
            damage_notes TEXT,
            recognition_notes TEXT,
            confidence TEXT,
            created_at TEXT NOT NULL
        )
    """)


    cur.execute("""
        CREATE TABLE IF NOT EXISTS remember_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token TEXT NOT NULL UNIQUE,
            expires_at TEXT NOT NULL,
            created_at TEXT NOT NULL,
            user_agent TEXT
        )
    """)

    conn.commit()

    safe_add_column("users", "annual_leave_entitlement", "REAL DEFAULT 28")

    conn.close()



def ensure_schema_updates():
    conn = get_db()
    cur = conn.cursor()

    additions = {
        "users": [
            ("role", "TEXT NOT NULL DEFAULT 'Admin'"),
            ("company_name", "TEXT"),
            ("stripe_customer_id", "TEXT"),
            ("stripe_subscription_id", "TEXT"),
            ("subscription_status", "TEXT DEFAULT 'manual'"),
            ("pro_expires_at", "TEXT"),
            ("pro_reason", "TEXT"),
            ("door_count", "INTEGER DEFAULT 100"),
            ("fence_count", "INTEGER DEFAULT 120"),
            ("door_start", "INTEGER DEFAULT 1"),
            ("door_end", "INTEGER DEFAULT 100"),
            ("fence_start", "INTEGER DEFAULT 1"),
            ("fence_end", "INTEGER DEFAULT 120")
        ],
        "team_members": [
            ("permissions", "TEXT DEFAULT 'View only'")
        ],
        "invoices": [
            ("email_sent", "INTEGER DEFAULT 0")
        ],
        "handovers": [
            ("pdf_created", "INTEGER DEFAULT 0")
        ],
        "photo_records": [
            ("ai_result", "TEXT")
        ]
    }

    for table, cols in additions.items():
        existing = [row["name"] for row in cur.execute(f"PRAGMA table_info({table})").fetchall()]
        for col_name, col_def in cols:
            if col_name not in existing:
                cur.execute(f"ALTER TABLE {table} ADD COLUMN {col_name} {col_def}")

    conn.commit()
    conn.close()




def seed_admin_user():
    conn = get_db()

    # Always make sure new columns exist before touching admin.
    existing_columns = [row["name"] for row in conn.execute("PRAGMA table_info(users)").fetchall()]

    def add_col(name, sql_type):
        nonlocal existing_columns
        if name not in existing_columns:
            conn.execute(f"ALTER TABLE users ADD COLUMN {name} {sql_type}")
            existing_columns.append(name)

    add_col("role", "TEXT NOT NULL DEFAULT 'Admin'")
    add_col("company_name", "TEXT")
    add_col("stripe_customer_id", "TEXT")
    add_col("stripe_subscription_id", "TEXT")
    add_col("subscription_status", "TEXT DEFAULT 'manual'")
    add_col("mileage_rate", "REAL DEFAULT 0.45")
    add_col("door_count", "INTEGER DEFAULT 100")
    add_col("fence_count", "INTEGER DEFAULT 120")
    add_col("door_start", "INTEGER DEFAULT 1")
    add_col("door_end", "INTEGER DEFAULT 100")
    add_col("fence_start", "INTEGER DEFAULT 1")
    add_col("fence_end", "INTEGER DEFAULT 120")
    add_col("pro_expires_at", "TEXT")
    add_col("pro_reason", "TEXT")
    conn.commit()

    existing = conn.execute("SELECT id FROM users WHERE email = ?", ("admin@opspilot.ai",)).fetchone()

    if not existing:
        conn.execute("""
            INSERT INTO users
            (name, email, password_hash, plan, business_name, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            "OpsPilot Admin",
            "admin@opspilot.ai",
            generate_password_hash("admin123"),
            "pro",
            "OpsPilot AI",
            datetime.now().isoformat()
        ))
        conn.commit()

    conn.execute("""
        UPDATE users
        SET plan = 'pro',
            role = 'Admin',
            company_name = 'OpsPilot AI Admin',
            subscription_status = 'admin',
            mileage_rate = 0.45,
            door_count = 100,
            door_start = 1,
            door_end = 100,
            fence_count = 120,
            fence_start = 1,
            fence_end = 120
        WHERE email = ?
    """, ("admin@opspilot.ai",))
    conn.commit()
    conn.close()

def is_admin(user=None):
    user = user or current_user()
    return bool(user and user["email"] == "admin@opspilot.ai")


def current_user():
    if "user_id" not in session:
        return None
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (session["user_id"],)).fetchone()
    conn.close()
    return refresh_user_plan(user)


@app.context_processor
def inject_context():
    user = current_user()
    return {
        "current_year": datetime.now().year,
        "user": user,
        "plan_names": PLAN_NAMES,
        "shift_trial": shift_calendar_trial_info(user) if user else None
    }




def get_system_setting(key, default=None):
    conn = get_db()
    row = conn.execute("SELECT value FROM system_settings WHERE key = ?", (key,)).fetchone()
    conn.close()
    return row["value"] if row and row["value"] is not None else default


def set_system_setting(key, value):
    conn = get_db()
    conn.execute("""
        INSERT INTO system_settings (key, value)
        VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
    """, (key, value))
    conn.commit()
    conn.close()


def get_stripe_config():
    return {
        "secret_key": get_system_setting("stripe_secret_key", os.environ.get("STRIPE_SECRET_KEY", "")),
        "publishable_key": get_system_setting("stripe_publishable_key", os.environ.get("STRIPE_PUBLISHABLE_KEY", "")),
        "webhook_secret": get_system_setting("stripe_webhook_secret", os.environ.get("STRIPE_WEBHOOK_SECRET", "")),
        "pro_price_id": get_system_setting("stripe_pro_price_id", os.environ.get("STRIPE_PRO_PRICE_ID", "")),
        "app_base_url": get_system_setting("app_base_url", os.environ.get("APP_BASE_URL", "http://127.0.0.1:5000")),
        "billing_mode": get_system_setting("billing_mode", "rolling"),
        "payout_note": get_system_setting("payout_note", "Stripe payouts are controlled from your Stripe Dashboard payout settings."),
    }


def first_day_next_month_timestamp():
    now = datetime.now()
    if now.month == 12:
        next_month = datetime(now.year + 1, 1, 1)
    else:
        next_month = datetime(now.year, now.month + 1, 1)
    return int(next_month.timestamp())


def stripe_ready():
    cfg = get_stripe_config()
    return bool(stripe and cfg["secret_key"] and cfg["pro_price_id"])


def row_get(row, key, default=None):
    try:
        if row is None:
            return default
        return row[key]
    except Exception:
        return default




def get_yard_config(user):
    is_pro_user = row_get(user, "plan") in ["pro", "business"] or is_admin(user)

    if not is_pro_user:
        return {
            "door_start": 1,
            "door_end": 100,
            "fence_start": 1,
            "fence_end": 120,
            "editable": False,
        }

    door_start = int(row_get(user, "door_start", 1) or 1)
    door_end = int(row_get(user, "door_end", row_get(user, "door_count", 100)) or 100)
    fence_start = int(row_get(user, "fence_start", 1) or 1)
    fence_end = int(row_get(user, "fence_end", row_get(user, "fence_count", 120)) or 120)

    door_start = max(1, min(door_start, 9999))
    door_end = max(door_start, min(door_end, 9999))
    fence_start = max(1, min(fence_start, 9999))
    fence_end = max(fence_start, min(fence_end, 9999))

    return {
        "door_start": door_start,
        "door_end": door_end,
        "fence_start": fence_start,
        "fence_end": fence_end,
        "editable": True,
    }


def refresh_user_plan(user):
    if not user:
        return user

    try:
        email = row_get(user, "email")
        expires_at = row_get(user, "pro_expires_at")

        if email == "admin@opspilot.ai":
            return user

        if expires_at:
            expiry = datetime.fromisoformat(expires_at)
            if datetime.now() > expiry:
                conn = get_db()
                conn.execute("""
                    UPDATE users
                    SET plan = 'free',
                        subscription_status = 'expired',
                        pro_expires_at = NULL,
                        pro_reason = NULL
                    WHERE id = ?
                """, (user["id"],))
                conn.commit()
                refreshed = conn.execute("SELECT * FROM users WHERE id = ?", (user["id"],)).fetchone()
                conn.close()
                return refreshed
    except Exception:
        return user

    return user



def shift_calendar_trial_info(user):
    """Free users can use Shift Calendar for 14 days after account creation."""
    if not user:
        return {"allowed": False, "days_left": 0, "is_trial": False}

    if is_admin(user) or row_get(user, "plan") in ["pro", "business"]:
        return {"allowed": True, "days_left": None, "is_trial": False}

    try:
        created = datetime.fromisoformat(row_get(user, "created_at"))
        trial_end = created + timedelta(days=14)
        now = datetime.now()
        days_left = max((trial_end.date() - now.date()).days, 0)
        return {
            "allowed": now <= trial_end,
            "days_left": days_left,
            "is_trial": True,
            "trial_end": trial_end.date().isoformat(),
        }
    except Exception:
        return {"allowed": False, "days_left": 0, "is_trial": True}



def create_remember_token(user_id):
    token = secrets.token_urlsafe(48)
    expires = datetime.now() + timedelta(days=30)
    conn = get_db()
    conn.execute("""
        INSERT INTO remember_tokens (user_id, token, expires_at, created_at, user_agent)
        VALUES (?, ?, ?, ?, ?)
    """, (user_id, token, expires.isoformat(), datetime.now().isoformat(), request.headers.get("User-Agent", "")))
    conn.commit()
    conn.close()
    return token, expires


def consume_remember_token():
    if session.get("user_id"):
        return

    token = request.cookies.get("opspilot_remember")
    if not token:
        return

    conn = get_db()
    row = conn.execute("""
        SELECT rt.*, u.id AS uid
        FROM remember_tokens rt
        JOIN users u ON u.id = rt.user_id
        WHERE rt.token = ?
    """, (token,)).fetchone()
    conn.close()

    if not row:
        return

    try:
        expires = datetime.fromisoformat(row["expires_at"])
        if expires < datetime.now():
            conn = get_db()
            conn.execute("DELETE FROM remember_tokens WHERE token = ?", (token,))
            conn.commit()
            conn.close()
            return
        session["user_id"] = row["user_id"]
    except Exception:
        return


def annual_leave_summary(user_id):
    user = current_user()
    entitlement = 28
    try:
        entitlement = float(row_get(user, "annual_leave_entitlement") or 28)
    except Exception:
        entitlement = 28

    year = datetime.today().year
    conn = get_db()
    used_rows = conn.execute("""
        SELECT COUNT(*) AS used
        FROM shift_calendar
        WHERE user_id = ?
          AND status = 'Holiday'
          AND date BETWEEN ? AND ?
    """, (user_id, f"{year}-01-01", f"{year}-12-31")).fetchone()
    conn.close()

    used = float(used_rows["used"] if used_rows else 0)
    return {
        "entitlement": entitlement,
        "used": used,
        "remaining": max(entitlement - used, 0),
        "year": year,
    }


def today_shift_status(user_id):
    today = datetime.today().date().isoformat()
    conn = get_db()
    row = conn.execute("""
        SELECT * FROM shift_calendar
        WHERE user_id = ? AND date = ?
        LIMIT 1
    """, (user_id, today)).fetchone()
    conn.close()

    if row:
        return row

    return {
        "date": today,
        "status": "Not Set",
        "shift_name": "",
        "start_time": "",
        "end_time": "",
        "notes": "",
        "source": "Empty",
    }


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            flash("Please login first.", "error")
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


def plan_required(required_plan):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user:
                flash("Please login first.", "error")
                return redirect(url_for("login"))
            if is_admin(user):
                return fn(*args, **kwargs)
            if PLAN_ORDER.get(user["plan"], 0) < PLAN_ORDER.get(required_plan, 0):
                flash(f"This feature requires {PLAN_NAMES[required_plan]} plan.", "error")
                return redirect(url_for("pricing"))
            return fn(*args, **kwargs)
        return wrapper
    return decorator


def money(value):
    return f"£{float(value or 0):,.2f}"


app.jinja_env.filters["money"] = money


@app.template_filter("dateuk")
def dateuk(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return value


def calculate_tax(profit):
    profit = float(profit or 0)
    basic_band = max(min(profit, BASIC_RATE_LIMIT) - PERSONAL_ALLOWANCE, 0)
    higher_band = max(profit - BASIC_RATE_LIMIT, 0)
    income_tax = basic_band * INCOME_TAX_BASIC + higher_band * INCOME_TAX_HIGHER

    class4_basic_band = max(min(profit, CLASS4_NI_UPPER) - CLASS4_NI_LOWER, 0)
    class4_higher_band = max(profit - CLASS4_NI_UPPER, 0)
    class4_ni = class4_basic_band * CLASS4_NI_BASIC + class4_higher_band * CLASS4_NI_HIGHER

    class2_ni = CLASS2_NI_WEEKLY * 52 if profit > 6725 else 0
    return income_tax + class4_ni + class2_ni


def totals(user_id):
    conn = get_db()
    income = conn.execute("SELECT COALESCE(SUM(amount), 0) FROM invoices WHERE user_id = ?", (user_id,)).fetchone()[0]
    expenses = conn.execute("SELECT COALESCE(SUM(amount), 0) FROM expenses WHERE user_id = ?", (user_id,)).fetchone()[0]
    miles = conn.execute("SELECT COALESCE(SUM(miles), 0) FROM mileage WHERE user_id = ?", (user_id,)).fetchone()[0]
    conn.close()

    mileage_claim = miles * HMRC_MILE_RATE
    taxable_profit = max(income - expenses - mileage_claim, 0)
    estimated_tax = calculate_tax(taxable_profit)
    net_profit = income - expenses

    return {
        "income": income,
        "expenses": expenses,
        "miles": miles,
        "mileage_claim": mileage_claim,
        "taxable_profit": taxable_profit,
        "estimated_tax": estimated_tax,
        "net_profit": net_profit,
    }


def next_invoice_number(user_id):
    conn = get_db()
    row = conn.execute("SELECT id FROM invoices WHERE user_id = ? ORDER BY id DESC LIMIT 1", (user_id,)).fetchone()
    conn.close()
    next_id = (row["id"] + 1) if row else 1
    return f"OP-{datetime.now().year}-{next_id:04d}"



def generate_morning_brief_text(date, shift, manager, role, volume, available_hc, late_trailers, safety_message, priorities, team_messages, break_reminder, equipment_reminder):
    return f"""Good morning team,

Today we are running {shift} shift.

Manager: {manager or 'Not specified'}
Role: {role or 'Not specified'}

Today's Plan:
- Expected volume: {volume}
- Available HC: {available_hc}
- Late trailers: {late_trailers}

Safety Message:
{safety_message or 'Keep the area safe, clean and controlled. Report hazards immediately.'}

Priorities:
{priorities or 'Focus on trailer control, clean handovers and completing work safely.'}

Team Messages:
{team_messages or 'Keep communication clear and support each other during the shift.'}

Break Reminder:
{break_reminder or 'Take breaks in a controlled way and make sure the operation is covered.'}

Equipment / MHE / Scanner / Key Reminder:
{equipment_reminder or 'Return all MHE keys, scanners and equipment at the end of the shift.'}

Let’s keep it safe, organised and productive.
"""

def generate_handover_text(date, shift, manager, volume, planned_hc, actual_hc, late_trailers, issues, actions):
    staffing_gap = actual_hc - planned_hc
    if staffing_gap < 0:
        staffing_line = f"Short by {abs(staffing_gap)} colleague(s) against plan."
    elif staffing_gap > 0:
        staffing_line = f"{staffing_gap} colleague(s) above plan."
    else:
        staffing_line = "Headcount matched the plan."

    return f"""Daily Handover Report

Date: {date}
Shift: {shift}
Manager: {manager or 'Not specified'}

Operational Summary:
- Volume handled: {volume}
- Planned HC: {planned_hc}
- Actual HC: {actual_hc}
- Late trailers: {late_trailers}
- Staffing: {staffing_line}

Issues:
{issues or 'No major issues reported.'}

Actions / Follow-up:
{actions or 'No further action required.'}

Next Shift Focus:
1. Confirm yard and trailer status early.
2. Review late trailers and unresolved issues.
3. Check headcount against expected volume.
4. Keep the shift handover clear and agreed before finish.
"""


def generate_shift_plan(date, shift, volume, available_hc, target_rate, planned_hours):
    capacity = available_hc * target_rate * planned_hours if available_hc and target_rate and planned_hours else 0
    gap = capacity - volume
    suggested_hc = volume / (target_rate * planned_hours) if target_rate and planned_hours else 0

    if capacity <= 0:
        status = "Not enough information to calculate capacity."
    elif gap >= 0:
        status = f"Plan looks achievable. Estimated spare capacity: {gap:,.0f} units."
    else:
        status = f"Risk: estimated shortfall of {abs(gap):,.0f} units. Consider extra labour, overtime or priority control."

    return f"""AI Shift Planner

Date: {date}
Shift: {shift}

Input:
- Expected volume: {volume}
- Available HC: {available_hc}
- Target rate/person/hour: {target_rate}
- Planned working hours: {planned_hours}

Capacity:
- Estimated capacity: {capacity:,.0f} units
- Suggested HC required: {suggested_hc:.1f}

Result:
{status}

Suggested Plan:
1. Start priority trailers/lanes first.
2. Put strongest colleagues in the highest risk area.
3. Review progress after the first break.
4. If behind plan, move labour early rather than waiting until end of shift.
5. Create a daily handover with remaining risks and actions.
"""



def get_openai_client():
    if not OPENAI_API_KEY or OpenAI is None:
        return None
    return OpenAI(api_key=OPENAI_API_KEY)


def generate_ai_shift_plan(date, shift, volume, available_hc, target_rate, planned_hours):
    fallback = generate_shift_plan(date, shift, volume, available_hc, target_rate, planned_hours)
    client = get_openai_client()
    if not client:
        return fallback + "\n\nAI Mode: fallback planner used. Add OPENAI_API_KEY for live AI planning."

    prompt = f"""
Create a practical warehouse shift plan.
Date: {date}
Shift: {shift}
Expected volume: {volume}
Available HC: {available_hc}
Target rate per person per hour: {target_rate}
Planned hours: {planned_hours}

Give:
- capacity estimate
- risk level
- staffing recommendation
- break review points
- manager actions
- handover notes
Use simple UK warehouse English.
"""
    try:
        response = client.chat.completions.create(
            model=os.environ.get("OPENAI_TEXT_MODEL", "gpt-4o-mini"),
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
        )
        return response.choices[0].message.content.strip()
    except Exception as exc:
        return fallback + f"\n\nAI Mode failed, fallback used. Error: {exc}"


def analyse_photo_ai(trailer_id, location_detail, damage_notes, recognition_notes):
    client = get_openai_client()
    if not client:
        return (
            "Photo AI fallback result:\\n"
            f"- Trailer ID entered: {trailer_id or 'Not provided'}\\n"
            f"- Location entered: {location_detail or 'Not provided'}\\n"
            f"- Damage notes: {damage_notes or 'None'}\\n"
            f"- Recognition notes: {recognition_notes or 'None'}\\n"
            "Add OPENAI_API_KEY and vision model integration for automatic image reading."
        )

    prompt = f"""
Review this manually entered trailer photo record and create a professional recognition summary.
Trailer ID: {trailer_id}
Location: {location_detail}
Damage notes: {damage_notes}
Recognition notes: {recognition_notes}

Return:
- trailer identification confidence
- possible issues
- recommended action
- handover note
"""
    try:
        response = client.chat.completions.create(
            model=os.environ.get("OPENAI_TEXT_MODEL", "gpt-4o-mini"),
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
        )
        return response.choices[0].message.content.strip()
    except Exception as exc:
        return f"Photo AI failed, manual record saved. Error: {exc}"


def create_handover_pdf(row):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("OpsPilot AI - Daily Handover Report", styles["Title"]))
    story.append(Spacer(1, 8))
    for line in row["generated_report"].splitlines():
        if line.strip():
            story.append(Paragraph(line.replace("&", "&amp;"), styles["BodyText"]))
        else:
            story.append(Spacer(1, 6))
    doc.build(story)
    buffer.seek(0)
    return buffer


def invoice_pdf_buffer(user, invoice):
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    pdf.setFont("Helvetica-Bold", 24)
    pdf.drawString(25 * mm, height - 30 * mm, "INVOICE")
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(25 * mm, height - 48 * mm, user["business_name"] or "OpsPilot AI Business")
    pdf.setFont("Helvetica", 10)

    y = height - 56 * mm
    for line in [user["name"], user["email"], user["phone"], user["address"]]:
        if line:
            pdf.drawString(25 * mm, y, str(line))
            y -= 6 * mm

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(130 * mm, height - 48 * mm, "Invoice No:")
    pdf.drawString(130 * mm, height - 56 * mm, "Date:")
    pdf.drawString(130 * mm, height - 64 * mm, "Status:")

    pdf.setFont("Helvetica", 11)
    pdf.drawString(158 * mm, height - 48 * mm, invoice["invoice_number"])
    pdf.drawString(158 * mm, height - 56 * mm, invoice["date"])
    pdf.drawString(158 * mm, height - 64 * mm, invoice["status"])

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(25 * mm, height - 90 * mm, "Bill To")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(25 * mm, height - 98 * mm, invoice["customer_name"])
    if invoice["customer_email"]:
        pdf.drawString(25 * mm, height - 106 * mm, invoice["customer_email"])

    pdf.line(25 * mm, height - 125 * mm, 185 * mm, height - 125 * mm)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(25 * mm, height - 134 * mm, "Description")
    pdf.drawRightString(185 * mm, height - 134 * mm, "Amount")
    pdf.line(25 * mm, height - 140 * mm, 185 * mm, height - 140 * mm)

    pdf.setFont("Helvetica", 11)
    pdf.drawString(25 * mm, height - 150 * mm, invoice["description"][:80])
    pdf.drawRightString(185 * mm, height - 150 * mm, f"£{invoice['amount']:,.2f}")

    pdf.line(25 * mm, height - 160 * mm, 185 * mm, height - 160 * mm)
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawRightString(160 * mm, height - 176 * mm, "Total:")
    pdf.drawRightString(185 * mm, height - 176 * mm, f"£{invoice['amount']:,.2f}")
    pdf.setFont("Helvetica", 9)
    pdf.drawString(25 * mm, 20 * mm, "Generated by OpsPilot AI.")
    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return buffer


def send_invoice_email(user, invoice):
    if not invoice["customer_email"]:
        return False, "Customer email is missing."
    if not all([SMTP_HOST, SMTP_USER, SMTP_PASSWORD, SMTP_FROM]):
        return False, "SMTP is not configured. Add SMTP_HOST, SMTP_USER, SMTP_PASSWORD and SMTP_FROM environment variables."

    msg = EmailMessage()
    msg["Subject"] = f"Invoice {invoice['invoice_number']} from {user['business_name'] or user['name']}"
    msg["From"] = SMTP_FROM
    msg["To"] = invoice["customer_email"]
    msg.set_content(f"""Hi {invoice['customer_name']},

Please find attached invoice {invoice['invoice_number']}.

Amount: £{invoice['amount']:,.2f}

Kind regards,
{user['name']}
""")

    pdf_data = invoice_pdf_buffer(user, invoice).read()
    msg.add_attachment(pdf_data, maintype="application", subtype="pdf", filename=f"{invoice['invoice_number']}.pdf")

    context = ssl.create_default_context()
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls(context=context)
        server.login(SMTP_USER, SMTP_PASSWORD)
        server.send_message(msg)

    return True, "Invoice email sent."




def calculate_google_maps_miles(origin, destination):
    if not GOOGLE_MAPS_API_KEY:
        return None, "Google Maps API key is missing. Add GOOGLE_MAPS_API_KEY to your environment variables."

    if requests is None:
        return None, "Requests package is missing."

    url = "https://maps.googleapis.com/maps/api/distancematrix/json"
    params = {
        "origins": origin,
        "destinations": destination,
        "units": "imperial",
        "mode": "driving",
        "key": GOOGLE_MAPS_API_KEY
    }

    try:
        response = requests.get(url, params=params, timeout=10)
        data = response.json()

        if data.get("status") != "OK":
            return None, f"Google Maps error: {data.get('status')}"

        element = data["rows"][0]["elements"][0]
        if element.get("status") != "OK":
            return None, f"Route error: {element.get('status')}"

        meters = element["distance"]["value"]
        miles = round(meters / 1609.344, 1)
        duration = element.get("duration", {}).get("text", "")
        distance_text = element.get("distance", {}).get("text", f"{miles} mi")

        return {
            "miles": miles,
            "duration": duration,
            "distance_text": distance_text,
            "origin": data.get("origin_addresses", [origin])[0],
            "destination": data.get("destination_addresses", [destination])[0],
            "claim": round(miles * HMRC_MILE_RATE, 2)
        }, None

    except Exception as exc:
        return None, f"Could not calculate mileage: {exc}"



SHIFT_STATUS_COLORS = {
    "Work": "work",
    "Off": "off",
    "Holiday": "holiday",
    "Sick": "sick",
    "Training": "training",
    "Overtime": "overtime",
    "Bank Holiday": "bankholiday",
    "Custom": "custom",
}


def parse_custom_shift_pattern(pattern_text):
    """
    Converts human-friendly pattern text to a repeating status cycle.
    Examples:
    - "3 on 4 off 4 on 3 off"
    - "2 work 2 off 3 holiday"
    - "4 day 4 off"
    """
    text = (pattern_text or "").lower().strip()
    if not text:
        return None

    text = text.replace("/", " ").replace(",", " ").replace("-", " ")
    text = text.replace("work", "on").replace("working", "on").replace("days", "on").replace("day", "on")
    text = text.replace("rest", "off").replace("offs", "off")
    text = text.replace("annual leave", "holiday")
    text = text.replace("sickness", "sick").replace("ill", "sick")
    text = re.sub(r"\s+", " ", text)

    tokens = text.split()
    cycle = []
    i = 0
    status_map = {
        "on": "Work",
        "off": "Off",
        "holiday": "Holiday",
        "sick": "Sick",
        "training": "Training",
        "overtime": "Overtime",
        "bankholiday": "Bank Holiday",
        "bank": "Bank Holiday",
        "custom": "Custom",
    }

    while i < len(tokens) - 1:
        if tokens[i].isdigit():
            count = int(tokens[i])
            word = tokens[i + 1]
            if word == "bank" and i + 2 < len(tokens) and tokens[i + 2] == "holiday":
                word = "bankholiday"
                i += 1
            status = status_map.get(word)
            if status and 0 < count <= 31:
                cycle.extend([status] * count)
                i += 2
                continue
        i += 1

    return cycle if cycle else None


def generate_shift_pattern_dates(start_date, pattern, months, shift_name, start_time, end_time, custom_pattern=""):
    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = start + timedelta(days=int(months or 12) * 31)
    rows = []

    if pattern == "custom":
        cycle = parse_custom_shift_pattern(custom_pattern) or (["Work"] * 4 + ["Off"] * 4)
    elif pattern == "4on4off":
        cycle = ["Work"] * 4 + ["Off"] * 4
    elif pattern == "5on2off":
        cycle = ["Work"] * 5 + ["Off"] * 2
    elif pattern == "monfri":
        cycle = None
    elif pattern == "2days2nights4off":
        cycle = ["Work"] * 4 + ["Off"] * 4
    elif pattern == "3on4off4on3off":
        cycle = ["Work"] * 3 + ["Off"] * 4 + ["Work"] * 4 + ["Off"] * 3
    else:
        cycle = ["Work"] * 4 + ["Off"] * 4

    d = start
    i = 0
    while d <= end:
        if pattern == "monfri":
            status = "Work" if d.weekday() < 5 else "Off"
        else:
            status = cycle[i % len(cycle)]

        rows.append({
            "date": d.isoformat(),
            "status": status,
            "shift_name": shift_name,
            "start_time": start_time if status in ["Work", "Training", "Overtime"] else "",
            "end_time": end_time if status in ["Work", "Training", "Overtime"] else "",
            "notes": custom_pattern if pattern == "custom" and i == 0 else "",
            "source": "Generated",
        })
        d += timedelta(days=1)
        i += 1

    return rows

def get_week_start(date_obj=None):
    date_obj = date_obj or datetime.today().date()
    return date_obj - timedelta(days=date_obj.weekday())


def get_current_week_shift_rows(user_id):
    week_start = get_week_start()
    week_end = week_start + timedelta(days=6)

    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM shift_calendar
        WHERE user_id = ? AND date BETWEEN ? AND ?
        ORDER BY date ASC
    """, (user_id, week_start.isoformat(), week_end.isoformat())).fetchall()
    conn.close()

    by_date = {row["date"]: row for row in rows}
    week = []
    for i in range(7):
        d = week_start + timedelta(days=i)
        key = d.isoformat()
        row = by_date.get(key)
        if row:
            week.append(row)
        else:
            week.append({
                "date": key,
                "status": "Not Set",
                "shift_name": "",
                "start_time": "",
                "end_time": "",
                "notes": "",
                "source": "Empty",
            })
    return week


def allowed_image(filename):
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext in {"png", "jpg", "jpeg", "webp"}



@app.before_request
def load_remembered_user():
    consume_remember_token()

@app.route("/")
@login_required
def index():
    user = current_user()
    conn = get_db()
    recent_mileage = conn.execute("SELECT * FROM mileage WHERE user_id = ? ORDER BY date DESC, id DESC LIMIT 5", (user["id"],)).fetchall()
    recent_yard = conn.execute("SELECT * FROM yard_checks WHERE user_id = ? ORDER BY date DESC, id DESC LIMIT 5", (user["id"],)).fetchall()
    conn.close()
    return render_template("dashboard.html", page="dashboard", totals=totals(user["id"]), recent_mileage=recent_mileage, recent_yard=recent_yard, week_shifts=get_current_week_shift_rows(user["id"]), today_shift=today_shift_status(user["id"]), annual_leave=annual_leave_summary(user["id"]))


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        if not name or not email or len(password) < 6:
            flash("Please enter name, email and a password with at least 6 characters.", "error")
            return redirect(url_for("register"))

        conn = get_db()
        try:
            cur = conn.execute("""
                INSERT INTO users (name, email, password_hash, plan, business_name, created_at)
                VALUES (?, ?, ?, 'free', ?, ?)
            """, (name, email, generate_password_hash(password), f"{name}'s Business", datetime.now().isoformat()))
            conn.commit()
            session["user_id"] = cur.lastrowid
        except sqlite3.IntegrityError:
            flash("Email already registered.", "error")
            conn.close()
            return redirect(url_for("register"))
        conn.close()
        return redirect(url_for("index"))
    return render_template("auth.html", mode="register", page="auth")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
        conn.close()
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            response = redirect(url_for("index"))
            if request.form.get("remember_me") == "yes":
                token, expires = create_remember_token(user["id"])
                response.set_cookie(
                    "opspilot_remember",
                    token,
                    max_age=60 * 60 * 24 * 30,
                    httponly=True,
                    secure=True,
                    samesite="Lax"
                )
            return response
        flash("Invalid login details.", "error")
    return render_template("auth.html", mode="login", page="auth")


@app.route("/logout")
def logout():
    session.clear()
    response = redirect(url_for("login"))
    response.delete_cookie("opspilot_remember")
    return response



@app.route("/shift-calendar", methods=["GET", "POST"])
@login_required
def shift_calendar():
    user = current_user()
    trial = shift_calendar_trial_info(user)
    if not trial["allowed"]:
        flash("Your 14 day free Shift Calendar trial has ended. Upgrade to Pro to continue using Shift Calendar.", "error")
        return redirect(url_for("pricing"))

    statuses = ["Work", "Off", "Holiday", "Sick", "Training", "Overtime", "Bank Holiday", "Custom"]
    patterns = [
        ("4on4off", "4 on / 4 off"),
        ("5on2off", "5 days / 2 off"),
        ("monfri", "Monday-Friday"),
        ("2days2nights4off", "2 days / 2 nights / 4 off"),
        ("3on4off4on3off", "3 on / 4 off / 4 on / 3 off"),
        ("custom", "Custom pattern"),
    ]

    if request.method == "POST":
        action = request.form.get("action")

        conn = get_db()

        if action == "generate":
            start_date = request.form.get("start_date") or datetime.today().strftime("%Y-%m-%d")
            pattern = request.form.get("pattern", "4on4off")
            months = int(request.form.get("months") or 12)
            shift_name = request.form.get("shift_name", "Shift").strip()
            start_time = request.form.get("start_time", "06:00")
            end_time = request.form.get("end_time", "18:00")
            replace_existing = request.form.get("replace_existing") == "yes"

            rows = generate_shift_pattern_dates(start_date, pattern, months, shift_name, start_time, end_time, request.form.get("custom_pattern", ""))

            if replace_existing:
                conn.execute("DELETE FROM shift_calendar WHERE user_id = ? AND date >= ?", (user["id"], start_date))

            for row in rows:
                conn.execute("""
                    INSERT INTO shift_calendar
                    (user_id, date, status, shift_name, start_time, end_time, notes, source, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(user_id, date) DO UPDATE SET
                        status = excluded.status,
                        shift_name = excluded.shift_name,
                        start_time = excluded.start_time,
                        end_time = excluded.end_time,
                        notes = excluded.notes,
                        source = excluded.source
                """, (
                    user["id"], row["date"], row["status"], row["shift_name"], row["start_time"],
                    row["end_time"], row["notes"], row["source"], datetime.now().isoformat()
                ))

            conn.commit()
            conn.close()
            flash(f"Shift calendar generated for {months} month(s).", "success")
            return redirect(url_for("shift_calendar"))

        if action == "manual":
            date = request.form.get("manual_date") or datetime.today().strftime("%Y-%m-%d")
            status = request.form.get("manual_status", "Work")
            shift_name = request.form.get("manual_shift_name", "").strip()
            start_time = request.form.get("manual_start_time", "")
            end_time = request.form.get("manual_end_time", "")
            notes = request.form.get("manual_notes", "").strip()

            conn.execute("""
                INSERT INTO shift_calendar
                (user_id, date, status, shift_name, start_time, end_time, notes, source, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'Manual', ?)
                ON CONFLICT(user_id, date) DO UPDATE SET
                    status = excluded.status,
                    shift_name = excluded.shift_name,
                    start_time = excluded.start_time,
                    end_time = excluded.end_time,
                    notes = excluded.notes,
                    source = 'Manual'
            """, (user["id"], date, status, shift_name, start_time, end_time, notes, datetime.now().isoformat()))

            conn.commit()
            conn.close()
            flash("Shift day updated.", "success")
            return redirect(url_for("shift_calendar"))

    start = request.args.get("start")
    try:
        start_date = datetime.strptime(start, "%Y-%m-%d").date() if start else get_week_start()
    except Exception:
        start_date = get_week_start()

    end_date = start_date + timedelta(days=27)

    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM shift_calendar
        WHERE user_id = ? AND date BETWEEN ? AND ?
        ORDER BY date ASC
    """, (user["id"], start_date.isoformat(), end_date.isoformat())).fetchall()
    conn.close()

    by_date = {row["date"]: row for row in rows}
    calendar_days = []
    d = start_date
    while d <= end_date:
        key = d.isoformat()
        row = by_date.get(key)
        calendar_days.append(row if row else {
            "date": key,
            "status": "Not Set",
            "shift_name": "",
            "start_time": "",
            "end_time": "",
            "notes": "",
            "source": "Empty",
        })
        d += timedelta(days=1)

    prev_start = (start_date - timedelta(days=28)).isoformat()
    next_start = (start_date + timedelta(days=28)).isoformat()

    return render_template(
        "shift_calendar.html",
        rows=calendar_days,
        statuses=statuses,
        patterns=patterns,
        page="shift_calendar",
        prev_start=prev_start,
        next_start=next_start,
        status_colors=SHIFT_STATUS_COLORS,
        trial=shift_calendar_trial_info(user),
    )


@app.route("/shift-calendar/export")
@login_required
@plan_required("pro")
def export_shift_calendar():
    user = current_user()
    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM shift_calendar
        WHERE user_id = ?
        ORDER BY date ASC
    """, (user["id"],)).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Shift Calendar"
    headers = ["Date", "Status", "Shift", "Start", "End", "Notes", "Source"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([
            row["date"], row["status"], row["shift_name"], row["start_time"],
            row["end_time"], row["notes"], row["source"]
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="shift-calendar.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/pricing")
@login_required
def pricing():
    return render_template("pricing.html", page="pricing")


@app.route("/set-plan/<plan>")
@login_required
def set_plan(plan):
    if plan not in PLAN_ORDER:
        flash("Invalid plan.", "error")
        return redirect(url_for("pricing"))
    user = current_user()
    conn = get_db()
    conn.execute("UPDATE users SET plan = ? WHERE id = ?", (plan, user["id"]))
    conn.commit()
    conn.close()
    flash(f"Demo plan changed to {PLAN_NAMES[plan]}. For real payments use Stripe checkout setup.", "success")
    return redirect(url_for("pricing"))



@app.route("/api/calculate-mileage", methods=["POST"])
@login_required
def api_calculate_mileage():
    data = request.get_json(silent=True) or {}
    origin = (data.get("origin") or "").strip()
    destination = (data.get("destination") or "").strip()

    if not origin or not destination:
        return {"ok": False, "error": "Please enter From and To postcode/address."}, 400

    result, error = calculate_google_maps_miles(origin, destination)
    if error:
        return {"ok": False, "error": error}, 400

    return {"ok": True, "result": result}


@app.route("/mileage", methods=["GET", "POST"])
@login_required
def mileage():
    user = current_user()
    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        start_location = request.form.get("start_location", "").strip()
        end_location = request.form.get("end_location", "").strip()
        purpose = request.form.get("purpose", "").strip()
        miles = float(request.form.get("miles") or 0)
        rate = float(request.form.get("rate") or user["mileage_rate"] or HMRC_MILE_RATE)
        if not start_location or not end_location or miles <= 0 or rate <= 0:
            flash("Please enter start, destination, miles and valid rate.", "error")
            return redirect(url_for("mileage"))
        conn = get_db()
        conn.execute("""
            INSERT INTO mileage (user_id, date, start_location, end_location, miles, rate, purpose, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (user["id"], date, start_location, end_location, miles, rate, purpose, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("Mileage added.", "success")
        return redirect(url_for("mileage"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM mileage WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()
    return render_template("mileage.html", rows=rows, rate=HMRC_MILE_RATE, default_rate=(user["mileage_rate"] or HMRC_MILE_RATE), page="mileage")



@app.route("/mileage/export")
@login_required
def export_mileage():
    user = current_user()
    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM mileage
        WHERE user_id = ?
        ORDER BY date DESC, id DESC
    """, (user["id"],)).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Mileage"

    headers = ["Date", "From", "To", "Purpose", "Miles", "Rate (£/mile)", "Claim (£)", "Recorded At"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    total_miles = 0
    total_claim = 0

    for row in rows:
        rate = row["rate"] if "rate" in row.keys() and row["rate"] else HMRC_MILE_RATE
        claim = float(row["miles"] or 0) * float(rate or 0)
        total_miles += float(row["miles"] or 0)
        total_claim += claim

        ws.append([
            row["date"],
            row["start_location"],
            row["end_location"],
            row["purpose"] or "",
            float(row["miles"] or 0),
            float(rate or 0),
            round(claim, 2),
            row["created_at"]
        ])

    ws.append([])
    ws.append(["TOTAL", "", "", "", round(total_miles, 2), "", round(total_claim, 2), ""])

    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"mileage-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/expenses", methods=["GET", "POST"])
@login_required
def expenses():
    user = current_user()
    categories = ["Fuel", "Parking", "Tolls", "Vehicle Maintenance", "Phone", "Equipment", "Insurance", "Office", "Other"]
    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        category = request.form.get("category", "Other")
        description = request.form.get("description", "").strip()
        amount = float(request.form.get("amount") or 0)
        if not description or amount <= 0:
            flash("Please enter description and amount.", "error")
            return redirect(url_for("expenses"))
        conn = get_db()
        conn.execute("""
            INSERT INTO expenses (user_id, date, category, description, amount, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (user["id"], date, category, description, amount, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("Expense added.", "success")
        return redirect(url_for("expenses"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM expenses WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()
    return render_template("expenses.html", rows=rows, categories=categories, page="expenses")



@app.route("/expenses/export")
@login_required
def export_expenses():
    user = current_user()
    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM expenses
        WHERE user_id = ?
        ORDER BY date DESC, id DESC
    """, (user["id"],)).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    headers = ["Date", "Category", "Description", "Amount (£)", "Recorded At"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    total_amount = 0

    for row in rows:
        amount = float(row["amount"] or 0)
        total_amount += amount
        ws.append([
            row["date"],
            row["category"],
            row["description"],
            round(amount, 2),
            row["created_at"]
        ])

    ws.append([])
    ws.append(["TOTAL", "", "", round(total_amount, 2), ""])

    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"expenses-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/invoices", methods=["GET", "POST"])
@login_required
def invoices():
    user = current_user()
    if request.method == "POST":
        invoice_number = request.form.get("invoice_number") or next_invoice_number(user["id"])
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        customer_name = request.form.get("customer_name", "").strip()
        customer_email = request.form.get("customer_email", "").strip()
        description = request.form.get("description", "").strip()
        amount = float(request.form.get("amount") or 0)
        status = request.form.get("status", "Unpaid")
        if not customer_name or not description or amount <= 0:
            flash("Please enter customer, description and amount.", "error")
            return redirect(url_for("invoices"))
        conn = get_db()
        conn.execute("""
            INSERT INTO invoices (user_id, invoice_number, date, customer_name, customer_email, description, amount, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user["id"], invoice_number, date, customer_name, customer_email, description, amount, status, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("Invoice created.", "success")
        return redirect(url_for("invoices"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM invoices WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()
    return render_template("invoices.html", rows=rows, invoice_number=next_invoice_number(user["id"]), page="invoices")


@app.route("/invoice/<int:item_id>/pdf")
@login_required
def invoice_pdf(item_id):
    user = current_user()
    conn = get_db()
    invoice = conn.execute("SELECT * FROM invoices WHERE id = ? AND user_id = ?", (item_id, user["id"])).fetchone()
    conn.close()
    if not invoice:
        flash("Invoice not found.", "error")
        return redirect(url_for("invoices"))

    buffer = invoice_pdf_buffer(user, invoice)
    return send_file(buffer, as_attachment=True, download_name=f"{invoice['invoice_number']}.pdf", mimetype="application/pdf")


@app.route("/invoice/<int:item_id>/email")
@login_required
def invoice_email(item_id):
    user = current_user()
    conn = get_db()
    invoice = conn.execute("SELECT * FROM invoices WHERE id = ? AND user_id = ?", (item_id, user["id"])).fetchone()
    if not invoice:
        conn.close()
        flash("Invoice not found.", "error")
        return redirect(url_for("invoices"))

    success, message = send_invoice_email(user, invoice)
    if success:
        conn.execute("UPDATE invoices SET email_sent = 1 WHERE id = ? AND user_id = ?", (item_id, user["id"]))
        conn.commit()
        flash(message, "success")
    else:
        flash(message, "error")
    conn.close()
    return redirect(url_for("invoices"))


@app.route("/tax")
@login_required
def tax():
    user = current_user()
    return render_template("tax.html", totals=totals(user["id"]), page="tax")



@app.route("/morning-brief", methods=["GET", "POST"])
@login_required
def morning_brief():
    user = current_user()

    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        shift = request.form.get("shift", "Day")
        manager = request.form.get("manager", "").strip()
        role = request.form.get("role", "").strip()
        volume = int(request.form.get("volume") or 0)
        available_hc = int(request.form.get("available_hc") or 0)
        late_trailers = int(request.form.get("late_trailers") or 0)
        safety_message = request.form.get("safety_message", "").strip()
        priorities = request.form.get("priorities", "").strip()
        team_messages = request.form.get("team_messages", "").strip()
        break_reminder = request.form.get("break_reminder", "").strip()
        equipment_reminder = request.form.get("equipment_reminder", "").strip()

        generated = generate_morning_brief_text(
            date, shift, manager, role, volume, available_hc, late_trailers,
            safety_message, priorities, team_messages, break_reminder, equipment_reminder
        )

        conn = get_db()
        conn.execute("""
            INSERT INTO morning_briefs
            (user_id, date, shift, manager, role, volume, available_hc, late_trailers,
             safety_message, priorities, team_messages, break_reminder, equipment_reminder, generated_brief, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            user["id"], date, shift, manager, role, volume, available_hc, late_trailers,
            safety_message, priorities, team_messages, break_reminder, equipment_reminder,
            generated, datetime.now().isoformat()
        ))
        conn.commit()
        conn.close()

        flash("Morning brief generated and saved.", "success")
        return redirect(url_for("morning_brief"))

    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM morning_briefs
        WHERE user_id = ?
        ORDER BY date DESC, id DESC
        LIMIT 50
    """, (user["id"],)).fetchall()
    conn.close()

    return render_template("morning_brief.html", rows=rows, page="morning_brief")


@app.route("/morning-brief/<int:item_id>/download")
@login_required
def morning_brief_download(item_id):
    user = current_user()
    conn = get_db()
    row = conn.execute("SELECT * FROM morning_briefs WHERE id = ? AND user_id = ?", (item_id, user["id"])).fetchone()
    conn.close()

    if not row:
        flash("Morning brief not found.", "error")
        return redirect(url_for("morning_brief"))

    output = BytesIO()
    output.write(row["generated_brief"].encode("utf-8"))
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=f"morning-brief-{row['date']}.txt", mimetype="text/plain")


@app.route("/yard-check", methods=["GET", "POST"])
@login_required
def yard_check():
    user = current_user()

    yard_cfg = get_yard_config(user)
    door_start = yard_cfg["door_start"]
    door_end = yard_cfg["door_end"]
    fence_start = yard_cfg["fence_start"]
    fence_end = yard_cfg["fence_end"]

    door_options = [f"Door {i}" for i in range(door_start, door_end + 1)]
    fence_options = [f"Fence {i}" for i in range(fence_start, fence_end + 1)]
    locations = ["Door", "Fence", "Yard", "Loading Bay", "Workshop", "Other"]
    statuses = ["Recorded", "Checked", "Issue Found", "Missing", "Moved", "Loaded", "Empty"]

    if request.method == "POST":
        if PLAN_ORDER[user["plan"]] < PLAN_ORDER["pro"] and not is_admin(user):
            flash("Free Yard Check demo: form works, but records are not saved. Upgrade to Pro to save and export.", "error")
            return redirect(url_for("yard_check"))

        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        entry_mode = request.form.get("entry_mode", "single")
        source = request.form.get("source", "Manual")
        status = request.form.get("status", "Recorded")
        notes = request.form.get("notes", "").strip()

        saved_count = 0
        conn = get_db()

        if entry_mode == "batch":
            batch_text = request.form.get("batch_text", "").strip()
            lines = [line.strip() for line in batch_text.splitlines() if line.strip()]

            for line in lines:
                original = line
                clean = line.replace("|", ",").replace(" at ", ",").replace(" AT ", ",")
                parts = [p.strip() for p in clean.split(",") if p.strip()]

                trailer_id = parts[0].upper() if parts else ""
                location_type = "Yard"
                location_detail = ""
                line_notes = notes or original

                if len(parts) >= 2:
                    loc = parts[1].title()
                    if loc.startswith("Door"):
                        location_type = "Door"
                        location_detail = loc
                    elif loc.startswith("Fence"):
                        location_type = "Fence"
                        location_detail = loc
                    else:
                        location_type = "Other"
                        location_detail = loc

                if len(parts) >= 3:
                    line_notes = parts[2]

                if trailer_id:
                    conn.execute("""
                        INSERT INTO yard_checks
                        (user_id, date, trailer_id, location_type, location_detail, status, notes, source, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (user["id"], date, trailer_id, location_type, location_detail, status, line_notes, source, datetime.now().isoformat()))
                    saved_count += 1

        else:
            yard_photo_filename = None
            yard_photo = request.files.get("photo") or request.files.get("camera_photo")
            if yard_photo and yard_photo.filename:
                if allowed_image(yard_photo.filename):
                    safe = secure_filename(yard_photo.filename)
                    yard_photo_filename = f"{uuid.uuid4().hex}_{safe}"
                    yard_photo.save(os.path.join(UPLOAD_DIR, yard_photo_filename))

            trailer_id = request.form.get("trailer_id", "").strip().upper()
            location_type = request.form.get("location_type", "Yard")
            location_detail = request.form.get("location_detail", "").strip()
            custom_location = request.form.get("custom_location", "").strip()

            if location_type == "Door":
                location_detail = request.form.get("door_number", location_detail)
            elif location_type == "Fence":
                location_detail = request.form.get("fence_number", location_detail)
            elif location_type == "Other" and custom_location:
                location_detail = custom_location

            if not trailer_id:
                conn.close()
                flash("Please enter trailer ID.", "error")
                return redirect(url_for("yard_check"))

            conn.execute("""
                INSERT INTO yard_checks
                (user_id, date, trailer_id, location_type, location_detail, status, notes, source, photo_filename, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (user["id"], date, trailer_id, location_type, location_detail, status, notes, source, yard_photo_filename, datetime.now().isoformat()))
            saved_count = 1

        conn.commit()
        conn.close()

        flash(f"{saved_count} trailer record(s) saved. Continue adding more.", "success")
        return redirect(url_for("yard_check"))

    search = request.args.get("search", "").strip()
    location_filter = request.args.get("location", "").strip()
    status_filter = request.args.get("status", "").strip()

    query = """
        SELECT * FROM yard_checks
        WHERE user_id = ?
    """
    params = [user["id"]]

    if search:
        like = f"%{search}%"
        query += """
            AND (
                trailer_id LIKE ?
                OR location_type LIKE ?
                OR location_detail LIKE ?
                OR status LIKE ?
                OR notes LIKE ?
                OR source LIKE ?
            )
        """
        params.extend([like, like, like, like, like, like])

    if location_filter:
        query += " AND location_type = ?"
        params.append(location_filter)

    if status_filter:
        query += " AND status = ?"
        params.append(status_filter)

    query += " ORDER BY date DESC, id DESC LIMIT 500"

    conn = get_db()
    rows = conn.execute(query, params).fetchall()

    today_count = conn.execute("""
        SELECT COUNT(*) FROM yard_checks
        WHERE user_id = ? AND date = ?
    """, (user["id"], datetime.today().strftime("%Y-%m-%d"))).fetchone()[0]

    door_used = conn.execute("""
        SELECT COUNT(*) FROM yard_checks
        WHERE user_id = ? AND location_type = 'Door'
    """, (user["id"],)).fetchone()[0]

    fence_used = conn.execute("""
        SELECT COUNT(*) FROM yard_checks
        WHERE user_id = ? AND location_type = 'Fence'
    """, (user["id"],)).fetchone()[0]

    conn.close()

    return render_template(
        "yard_check.html",
        rows=rows,
        locations=locations,
        statuses=statuses,
        door_options=door_options,
        fence_options=fence_options,
        door_start=door_start,
        door_end=door_end,
        fence_start=fence_start,
        fence_end=fence_end,
        yard_config_editable=yard_cfg["editable"],
        door_used=door_used,
        fence_used=fence_used,
        today_count=today_count,
        search=search,
        location_filter=location_filter,
        status_filter=status_filter,
        page="yard_check"
    )


@app.route("/yard-check/<int:item_id>/edit", methods=["GET", "POST"])
@login_required
@plan_required("pro")
def edit_yard_check(item_id):
    user = current_user()
    statuses = ["Recorded", "Checked", "Issue Found", "Missing", "Moved", "Loaded", "Empty"]
    locations = ["Door", "Fence", "Yard", "Loading Bay", "Workshop", "Other"]

    conn = get_db()
    row = conn.execute("SELECT * FROM yard_checks WHERE id = ? AND user_id = ?", (item_id, user["id"])).fetchone()

    if not row:
        conn.close()
        flash("Yard check not found.", "error")
        return redirect(url_for("yard_check"))

    if request.method == "POST":
        trailer_id = request.form.get("trailer_id", "").strip().upper()
        if not trailer_id:
            conn.close()
            flash("Trailer ID is required.", "error")
            return redirect(url_for("edit_yard_check", item_id=item_id))

        conn.execute("""
            UPDATE yard_checks
            SET date = ?, trailer_id = ?, location_type = ?, location_detail = ?, status = ?, notes = ?, source = ?
            WHERE id = ? AND user_id = ?
        """, (
            request.form.get("date") or row["date"],
            trailer_id,
            request.form.get("location_type", "Yard"),
            request.form.get("location_detail", "").strip(),
            request.form.get("status", "Recorded"),
            request.form.get("notes", "").strip(),
            request.form.get("source", "Manual"),
            item_id,
            user["id"]
        ))
        conn.commit()
        conn.close()
        flash("Yard check updated.", "success")
        return redirect(url_for("yard_check"))

    conn.close()
    return render_template("yard_edit.html", row=row, statuses=statuses, locations=locations, page="yard_check")


@app.route("/yard-check/<int:item_id>/delete", methods=["POST"])
@login_required
@plan_required("pro")
def delete_yard_check(item_id):
    user = current_user()
    conn = get_db()
    conn.execute("DELETE FROM yard_checks WHERE id = ? AND user_id = ?", (item_id, user["id"]))
    conn.commit()
    conn.close()
    flash("Yard check deleted.", "success")
    return redirect(url_for("yard_check"))



@app.route("/yard-check/edit/<int:item_id>", methods=["GET", "POST"])
@login_required
@plan_required("pro")
def edit_yard_check_old_url(item_id):
    return edit_yard_check(item_id)


@app.route("/yard-check/delete/<int:item_id>", methods=["POST"])
@login_required
@plan_required("pro")
def delete_yard_check_old_url(item_id):
    return delete_yard_check(item_id)


@app.route("/yard-check/export")
@login_required
@plan_required("pro")
def export_yard_check():
    user = current_user()
    conn = get_db()
    rows = conn.execute("SELECT * FROM yard_checks WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Yard Check"
    headers = ["Date", "Trailer ID", "Location", "Detail", "Status", "Notes", "Source", "Recorded At"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row["date"], row["trailer_id"], row["location_type"], row["location_detail"], row["status"], row["notes"], row["source"], row["created_at"]])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="yard-check.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/kpi", methods=["GET", "POST"])
@login_required
@plan_required("pro")
def kpi_dashboard():
    user = current_user()
    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        shift = request.form.get("shift", "Day")
        volume = int(request.form.get("volume") or 0)
        planned_hc = int(request.form.get("planned_hc") or 0)
        actual_hc = int(request.form.get("actual_hc") or 0)
        target_rate = float(request.form.get("target_rate") or 0)
        actual_rate = float(request.form.get("actual_rate") or 0)
        late_trailers = int(request.form.get("late_trailers") or 0)
        errors = int(request.form.get("errors") or 0)
        notes = request.form.get("notes", "").strip()
        conn = get_db()
        conn.execute("""
            INSERT INTO kpi_records (user_id, date, shift, volume, planned_hc, actual_hc, target_rate, actual_rate, late_trailers, errors, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user["id"], date, shift, volume, planned_hc, actual_hc, target_rate, actual_rate, late_trailers, errors, notes, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("KPI saved.", "success")
        return redirect(url_for("kpi_dashboard"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM kpi_records WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    summary = conn.execute("""
        SELECT COALESCE(SUM(volume), 0) total_volume, COALESCE(AVG(actual_rate), 0) avg_rate,
               COALESCE(SUM(late_trailers), 0) total_late, COALESCE(SUM(errors), 0) total_errors
        FROM kpi_records WHERE user_id = ?
    """, (user["id"],)).fetchone()
    conn.close()
    return render_template("kpi.html", rows=rows, summary=summary, page="kpi")


@app.route("/handover", methods=["GET", "POST"])
@login_required
@plan_required("pro")
def handover():
    user = current_user()
    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        shift = request.form.get("shift", "Day")
        manager = request.form.get("manager", "").strip()
        volume = int(request.form.get("volume") or 0)
        planned_hc = int(request.form.get("planned_hc") or 0)
        actual_hc = int(request.form.get("actual_hc") or 0)
        late_trailers = int(request.form.get("late_trailers") or 0)
        issues = request.form.get("issues", "").strip()
        actions = request.form.get("actions", "").strip()
        generated = generate_handover_text(date, shift, manager, volume, planned_hc, actual_hc, late_trailers, issues, actions)
        conn = get_db()
        conn.execute("""
            INSERT INTO handovers (user_id, date, shift, manager, volume, planned_hc, actual_hc, late_trailers, issues, actions, generated_report, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user["id"], date, shift, manager, volume, planned_hc, actual_hc, late_trailers, issues, actions, generated, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("Handover generated.", "success")
        return redirect(url_for("handover"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM handovers WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()
    return render_template("handover.html", rows=rows, page="handover")


@app.route("/handover/<int:item_id>/download")
@login_required
@plan_required("pro")
def handover_download(item_id):
    user = current_user()
    conn = get_db()
    row = conn.execute("SELECT * FROM handovers WHERE id = ? AND user_id = ?", (item_id, user["id"])).fetchone()
    conn.close()
    if not row:
        flash("Handover not found.", "error")
        return redirect(url_for("handover"))
    output = BytesIO()
    output.write(row["generated_report"].encode("utf-8"))
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"handover-{row['date']}.txt", mimetype="text/plain")




@app.route("/handover/<int:item_id>/pdf")
@login_required
@plan_required("pro")
def handover_pdf(item_id):
    user = current_user()
    conn = get_db()
    row = conn.execute("SELECT * FROM handovers WHERE id = ? AND user_id = ?", (item_id, user["id"])).fetchone()
    if not row:
        conn.close()
        flash("Handover not found.", "error")
        return redirect(url_for("handover"))
    conn.execute("UPDATE handovers SET pdf_created = 1 WHERE id = ? AND user_id = ?", (item_id, user["id"]))
    conn.commit()
    conn.close()
    buffer = create_handover_pdf(row)
    return send_file(buffer, as_attachment=True, download_name=f"handover-{row['date']}.pdf", mimetype="application/pdf")


@app.route("/team", methods=["GET", "POST"])
@login_required
@plan_required("business")
def team():
    user = current_user()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        role = request.form.get("role", "").strip()
        email = request.form.get("email", "").strip()
        status = request.form.get("status", "Active")
        permissions = request.form.get("permissions", "View only")
        if not name or not role:
            flash("Please enter name and role.", "error")
            return redirect(url_for("team"))
        conn = get_db()
        conn.execute("INSERT INTO team_members (user_id, name, role, email, status, permissions, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                     (user["id"], name, role, email, status, permissions, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("Team member added.", "success")
        return redirect(url_for("team"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM team_members WHERE user_id = ? ORDER BY name ASC", (user["id"],)).fetchall()
    conn.close()
    return render_template("team.html", rows=rows, page="team")


@app.route("/operations")
@login_required
@plan_required("business")
def operations():
    user = current_user()
    conn = get_db()
    kpi = conn.execute("""
        SELECT COALESCE(SUM(volume), 0) total_volume, COALESCE(AVG(actual_rate), 0) avg_rate,
               COALESCE(SUM(late_trailers), 0) total_late, COALESCE(SUM(errors), 0) total_errors
        FROM kpi_records WHERE user_id = ?
    """, (user["id"],)).fetchone()
    yard_count = conn.execute("SELECT COUNT(*) FROM yard_checks WHERE user_id = ?", (user["id"],)).fetchone()[0]
    team_count = conn.execute("SELECT COUNT(*) FROM team_members WHERE user_id = ?", (user["id"],)).fetchone()[0]
    handover_count = conn.execute("SELECT COUNT(*) FROM handovers WHERE user_id = ?", (user["id"],)).fetchone()[0]
    photo_count = conn.execute("SELECT COUNT(*) FROM photo_records WHERE user_id = ?", (user["id"],)).fetchone()[0]
    conn.close()
    return render_template("operations.html", kpi=kpi, yard_count=yard_count, team_count=team_count, handover_count=handover_count, photo_count=photo_count, page="operations")


@app.route("/shift-planner", methods=["GET", "POST"])
@login_required
@plan_required("business")
def shift_planner():
    user = current_user()
    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        shift = request.form.get("shift", "Day")
        volume = int(request.form.get("volume") or 0)
        available_hc = int(request.form.get("available_hc") or 0)
        target_rate = float(request.form.get("target_rate") or 0)
        planned_hours = float(request.form.get("planned_hours") or 0)
        ai_plan = generate_ai_shift_plan(date, shift, volume, available_hc, target_rate, planned_hours)
        conn = get_db()
        conn.execute("""
            INSERT INTO shift_plans (user_id, date, shift, volume, available_hc, target_rate, planned_hours, ai_plan, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user["id"], date, shift, volume, available_hc, target_rate, planned_hours, ai_plan, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("AI shift plan generated.", "success")
        return redirect(url_for("shift_planner"))
    conn = get_db()
    rows = conn.execute("SELECT * FROM shift_plans WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()
    return render_template("shift_planner.html", rows=rows, page="shift_planner")


@app.route("/photo-recognition", methods=["GET", "POST"])
@login_required
@plan_required("business")
def photo_recognition():
    user = current_user()
    if request.method == "POST":
        date = request.form.get("date") or datetime.today().strftime("%Y-%m-%d")
        trailer_id = request.form.get("trailer_id", "").strip().upper()
        location_detail = request.form.get("location_detail", "").strip()
        damage_notes = request.form.get("damage_notes", "").strip()
        recognition_notes = request.form.get("recognition_notes", "").strip()
        confidence = request.form.get("confidence", "Manual check")
        ai_result = analyse_photo_ai(trailer_id, location_detail, damage_notes, recognition_notes)
        image_filename = None

        uploaded = request.files.get("image")
        if uploaded and uploaded.filename:
            if not allowed_image(uploaded.filename):
                flash("Only PNG, JPG, JPEG or WEBP images are allowed.", "error")
                return redirect(url_for("photo_recognition"))
            safe = secure_filename(uploaded.filename)
            image_filename = f"{uuid.uuid4().hex}_{safe}"
            uploaded.save(os.path.join(UPLOAD_DIR, image_filename))

        conn = get_db()
        conn.execute("""
            INSERT INTO photo_records (user_id, date, image_filename, trailer_id, location_detail, damage_notes, recognition_notes, confidence, ai_result, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user["id"], date, image_filename, trailer_id, location_detail, damage_notes, recognition_notes, confidence, ai_result, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        flash("Photo recognition record saved.", "success")
        return redirect(url_for("photo_recognition"))

    conn = get_db()
    rows = conn.execute("SELECT * FROM photo_records WHERE user_id = ? ORDER BY date DESC, id DESC", (user["id"],)).fetchall()
    conn.close()
    return render_template("photo_recognition.html", rows=rows, page="photo_recognition")



@app.route("/billing/checkout/<plan>")
@login_required
def billing_checkout(plan):
    if plan not in ["pro", "business"]:
        flash("Invalid billing plan.", "error")
        return redirect(url_for("pricing"))

    user = current_user()

    if not stripe or not STRIPE_SECRET_KEY:
        flash("Stripe is not configured yet. For now, plan selection works in demo/manual mode.", "error")
        return redirect(url_for("pricing"))

    price_id = STRIPE_PRO_PRICE_ID if plan == "pro" else STRIPE_BUSINESS_PRICE_ID
    if not price_id:
        flash("Stripe price ID is missing for this plan.", "error")
        return redirect(url_for("pricing"))

    stripe.api_key = STRIPE_SECRET_KEY
    checkout = stripe.checkout.Session.create(
        mode="subscription",
        payment_method_types=["card"],
        customer_email=user["email"],
        line_items=[{"price": price_id, "quantity": 1}],
        success_url=f"{APP_BASE_URL}/billing/success/{plan}",
        cancel_url=f"{APP_BASE_URL}/pricing",
    )
    return redirect(checkout.url, code=303)


@app.route("/billing/success/<plan>")
@login_required
def billing_success(plan):
    if plan not in PLAN_ORDER:
        flash("Invalid plan.", "error")
        return redirect(url_for("pricing"))
    user = current_user()
    conn = get_db()
    conn.execute("UPDATE users SET plan = ?, subscription_status = 'active' WHERE id = ?", (plan, user["id"]))
    conn.commit()
    conn.close()
    flash(f"Subscription activated: {PLAN_NAMES[plan]}.", "success")
    return redirect(url_for("pricing"))



@app.route("/admin")
@login_required
def admin_dashboard():
    user = current_user()
    if not is_admin(user):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    conn = get_db()
    users = conn.execute("SELECT * FROM users ORDER BY created_at DESC").fetchall()

    stats = {
        "users": conn.execute("SELECT COUNT(*) FROM users").fetchone()[0],
        "free": conn.execute("SELECT COUNT(*) FROM users WHERE plan = 'free'").fetchone()[0],
        "pro": conn.execute("SELECT COUNT(*) FROM users WHERE plan = 'pro'").fetchone()[0],
        "business": conn.execute("SELECT COUNT(*) FROM users WHERE plan = 'business'").fetchone()[0],
        "yard": conn.execute("SELECT COUNT(*) FROM yard_checks").fetchone()[0],
        "kpi": conn.execute("SELECT COUNT(*) FROM kpi_records").fetchone()[0],
        "handovers": conn.execute("SELECT COUNT(*) FROM handovers").fetchone()[0],
        "team": conn.execute("SELECT COUNT(*) FROM team_members").fetchone()[0],
    }
    conn.close()

    return render_template("admin.html", users=users, stats=stats, page="admin")


@app.route("/admin/user/<int:user_id>/plan/<plan>")
@login_required
def admin_set_user_plan(user_id, plan):
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    if plan not in ["free", "pro"]:
        flash("Invalid plan.", "error")
        return redirect(url_for("admin_dashboard"))

    conn = get_db()

    if plan == "pro":
        conn.execute("""
            UPDATE users
            SET plan = 'pro',
                subscription_status = 'manual_admin',
                pro_expires_at = NULL,
                pro_reason = 'Manual Pro set by admin'
            WHERE id = ?
        """, (user_id,))
    else:
        conn.execute("""
            UPDATE users
            SET plan = 'free',
                subscription_status = 'manual_admin',
                pro_expires_at = NULL,
                pro_reason = NULL
            WHERE id = ?
        """, (user_id,))

    conn.commit()
    conn.close()

    flash(f"User plan changed to {PLAN_NAMES[plan]}.", "success")
    return redirect(url_for("admin_dashboard"))



@app.route("/admin/user/<int:user_id>/trial/pro30")
@login_required
def admin_set_user_trial_pro30(user_id):
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    expires_at = (datetime.now() + timedelta(days=30)).isoformat(timespec="seconds")

    conn = get_db()
    conn.execute("""
        UPDATE users
        SET plan = 'pro',
            subscription_status = 'trial_admin',
            pro_expires_at = ?,
            pro_reason = '30 day Pro trial set by admin'
        WHERE id = ?
    """, (expires_at, user_id))
    conn.commit()
    conn.close()

    flash("30 day Pro trial activated for user.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/user/<int:user_id>/gift/pro")
@login_required
def admin_set_user_gift_pro(user_id):
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    conn = get_db()
    conn.execute("""
        UPDATE users
        SET plan = 'pro',
            subscription_status = 'gift_admin',
            pro_expires_at = NULL,
            pro_reason = 'Free Pro gift set by admin'
        WHERE id = ?
    """, (user_id,))
    conn.commit()
    conn.close()

    flash("Free Pro gift activated for user.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/user/<int:user_id>/role/<role>")
@login_required
def admin_set_user_role(user_id, role):
    admin = current_user()
    if not is_admin(admin):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    allowed = ["Admin", "Manager", "FLM", "User"]
    if role not in allowed:
        flash("Invalid role.", "error")
        return redirect(url_for("admin_dashboard"))

    conn = get_db()
    conn.execute("UPDATE users SET role = ? WHERE id = ?", (role, user_id))
    conn.commit()
    conn.close()

    flash(f"User role changed to {role}.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/billing/checkout/pro")
@login_required
def billing_checkout_pro():
    user = current_user()
    cfg = get_stripe_config()

    if not stripe or not cfg["secret_key"] or not cfg["pro_price_id"]:
        flash("Stripe is not configured yet. Admin must add Stripe keys and Pro Price ID.", "error")
        return redirect(url_for("pricing"))

    stripe.api_key = cfg["secret_key"]

    subscription_data = {
        "metadata": {
            "user_id": str(user["id"]),
            "plan": "pro",
        }
    }

    # rolling = normal monthly subscription from today.
    # calendar_prorated = first invoice is proportional until the first day of next month,
    # then billing renews on the 1st monthly.
    if cfg.get("billing_mode") == "calendar_prorated":
        subscription_data["billing_cycle_anchor"] = first_day_next_month_timestamp()
        subscription_data["proration_behavior"] = "create_prorations"

    session = stripe.checkout.Session.create(
        mode="subscription",
        customer_email=user["email"],
        line_items=[{"price": cfg["pro_price_id"], "quantity": 1}],
        subscription_data=subscription_data,
        success_url=f"{cfg['app_base_url']}/billing/success?session_id={{CHECKOUT_SESSION_ID}}",
        cancel_url=f"{cfg['app_base_url']}/pricing",
        metadata={
            "user_id": str(user["id"]),
            "plan": "pro",
        },
    )

    return redirect(session.url, code=303)


@app.route("/billing/success")
@login_required
def billing_success_legacy_2():
    flash("Payment completed. If your plan has not updated yet, it will update after Stripe webhook confirmation.", "success")
    return redirect(url_for("pricing"))


@app.route("/billing/portal")
@login_required
def billing_portal():
    user = current_user()
    cfg = get_stripe_config()

    if not stripe or not cfg["secret_key"]:
        flash("Stripe is not configured yet.", "error")
        return redirect(url_for("pricing"))

    if not row_get(user, "stripe_customer_id"):
        flash("No Stripe customer found for this account yet.", "error")
        return redirect(url_for("pricing"))

    stripe.api_key = cfg["secret_key"]

    portal = stripe.billing_portal.Session.create(
        customer=user["stripe_customer_id"],
        return_url=f"{cfg['app_base_url']}/pricing",
    )

    return redirect(portal.url, code=303)


@app.route("/stripe/webhook", methods=["POST"])
def stripe_webhook():
    cfg = get_stripe_config()

    if not stripe or not cfg["webhook_secret"]:
        return "Stripe webhook not configured", 400

    payload = request.get_data()
    sig_header = request.headers.get("Stripe-Signature")

    try:
        event = stripe.Webhook.construct_event(payload, sig_header, cfg["webhook_secret"])
    except Exception:
        return "Invalid webhook", 400

    event_type = event["type"]
    obj = event["data"]["object"]

    conn = get_db()

    if event_type == "checkout.session.completed":
        user_id = obj.get("metadata", {}).get("user_id")
        customer_id = obj.get("customer")
        subscription_id = obj.get("subscription")

        if user_id:
            conn.execute("""
                UPDATE users
                SET plan = 'pro',
                    stripe_customer_id = ?,
                    stripe_subscription_id = ?,
                    subscription_status = 'active'
                WHERE id = ?
            """, (customer_id, subscription_id, user_id))
            conn.commit()

    elif event_type in ["invoice.payment_succeeded", "customer.subscription.updated"]:
        subscription_id = obj.get("subscription") or obj.get("id")
        status = obj.get("status", "active")

        if subscription_id:
            plan = "pro" if status in ["active", "trialing"] else "free"
            conn.execute("""
                UPDATE users
                SET plan = ?, subscription_status = ?
                WHERE stripe_subscription_id = ?
            """, (plan, status, subscription_id))
            conn.commit()

    elif event_type in ["customer.subscription.deleted", "customer.subscription.paused"]:
        subscription_id = obj.get("id")
        if subscription_id:
            conn.execute("""
                UPDATE users
                SET plan = 'free', subscription_status = ?
                WHERE stripe_subscription_id = ?
            """, (obj.get("status", "canceled"), subscription_id))
            conn.commit()

    conn.close()
    return "ok", 200



@app.route("/admin/billing", methods=["GET", "POST"])
@login_required
def admin_billing():
    user = current_user()
    if not is_admin(user):
        flash("Admin access only.", "error")
        return redirect(url_for("index"))

    if request.method == "POST":
        fields = [
            "stripe_publishable_key",
            "stripe_secret_key",
            "stripe_webhook_secret",
            "stripe_pro_price_id",
            "app_base_url",
            "billing_mode",
            "payout_note",
        ]
        for field in fields:
            set_system_setting(field, request.form.get(field, "").strip())

        flash("Billing settings saved.", "success")
        return redirect(url_for("admin_billing"))

    cfg = get_stripe_config()
    return render_template("admin_billing.html", cfg=cfg, page="admin")


@app.route("/settings/yard", methods=["GET", "POST"])
@login_required
@plan_required("pro")
def yard_settings():
    user = current_user()

    if request.method == "POST":
        door_start = int(request.form.get("door_start") or 1)
        door_end = int(request.form.get("door_end") or 100)
        fence_start = int(request.form.get("fence_start") or 1)
        fence_end = int(request.form.get("fence_end") or 120)

        door_start = max(1, min(door_start, 9999))
        door_end = max(door_start, min(door_end, 9999))
        fence_start = max(1, min(fence_start, 9999))
        fence_end = max(fence_start, min(fence_end, 9999))

        conn = get_db()
        conn.execute("""
            UPDATE users
            SET door_start = ?,
                door_end = ?,
                door_count = ?,
                fence_start = ?,
                fence_end = ?,
                fence_count = ?
            WHERE id = ?
        """, (
            door_start,
            door_end,
            door_end,
            fence_start,
            fence_end,
            fence_end,
            user["id"]
        ))
        conn.commit()
        conn.close()

        flash("Yard configuration saved.", "success")
        return redirect(url_for("yard_settings"))

    cfg = get_yard_config(user)
    return render_template("yard_settings.html", cfg=cfg, page="yard_settings")



@app.route("/settings/annual-leave", methods=["POST"])
@login_required
def save_annual_leave_settings():
    user = current_user()
    entitlement = request.form.get("annual_leave_entitlement", "28")
    try:
        entitlement = float(entitlement)
    except Exception:
        entitlement = 28

    conn = get_db()
    conn.execute("UPDATE users SET annual_leave_entitlement = ? WHERE id = ?", (entitlement, user["id"]))
    conn.commit()
    conn.close()

    flash("Annual leave entitlement saved.", "success")
    return redirect(url_for("settings"))

@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    user = current_user()
    if request.method == "POST":
        conn = get_db()
        conn.execute("""
            UPDATE users SET business_name = ?, company_name = ?, name = ?, role = ?, phone = ?, address = ?, mileage_rate = ?, door_count = ?, fence_count = ? WHERE id = ?
        """, (
            request.form.get("business_name", "").strip(),
            request.form.get("company_name", "").strip(),
            request.form.get("name", "").strip(),
            request.form.get("role", "Admin").strip(),
            request.form.get("phone", "").strip(),
            request.form.get("address", "").strip(),
            float(request.form.get("mileage_rate") or 0.45),
            int(request.form.get("door_count") or 100),
            int(request.form.get("fence_count") or 120),
            user["id"]
        ))
        conn.commit()
        conn.close()
        flash("Settings saved.", "success")
        return redirect(url_for("settings"))
    return render_template("settings.html", row=user, page="settings")



@app.route("/manifest.json")
def manifest_json():
    manifest = {
        "name": "OpsPilot AI",
        "short_name": "OpsPilot",
        "description": "Simple operations hub for managers.",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#f5f7fb",
        "theme_color": "#2563eb",
        "orientation": "portrait-primary",
        "icons": [
            {"src": "/static/icons/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/icons/icon-512.png", "sizes": "512x512", "type": "image/png"}
        ]
    }
    return app.response_class(json.dumps(manifest), mimetype="application/manifest+json")


@app.route("/service-worker.js")
def service_worker():
    js = """
const CACHE_NAME = 'opspilot-ai-v1';
const urlsToCache = [
  '/',
  '/static/css/style.css'
];

self.addEventListener('install', event => {
  event.waitUntil(
    caches.open(CACHE_NAME).then(cache => cache.addAll(urlsToCache)).catch(() => null)
  );
});

self.addEventListener('fetch', event => {
  event.respondWith(
    fetch(event.request).catch(() => caches.match(event.request))
  );
});
"""
    return app.response_class(js, mimetype="application/javascript")



init_db()
ensure_schema_updates()
seed_admin_user()
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
